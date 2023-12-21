# Import Statements
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import *
import ttkbootstrap as ttkb
from PIL import ImageTk
import webbrowser
import random
import csv
import openpyxl
import xlrd
import os
import sys

# Directories
home_directory = os.path.expanduser('~')
script_directory = os.path.dirname(sys.argv[0])
logo_icon_path = os.path.join(script_directory, 'logo.ico').replace('\\', '\\\\')
logo_img_path = os.path.join(script_directory, 'Images', 'logo.png').replace('\\', '\\\\')
import_formats_img_path = os.path.join(script_directory, 'Images', 'import.png').replace('\\', '\\\\')
key_img_path = os.path.join(script_directory, 'Images', 'key_format.png').replace('\\', '\\\\')
db_img_path = os.path.join(script_directory, 'Images', 'rename_db.png').replace('\\', '\\\\')
files_img_path = os.path.join(script_directory, 'Images', 'rename_files.png').replace('\\', '\\\\')
ideal_db_img_path = os.path.join(script_directory, 'Images', 'ideal_db.png').replace('\\', '\\\\')


def open_github():
    webbrowser.open('https://github.com/AlexHenriques/IDBlindingTool')


# IDBlindinTool: Original Labels
og_labels_list = [
    'brunswick green', 'fluorescent cyan', 'umber', 'pine green', 'caput mortuum', 'platinum', 'peach', 'red brown',
    'wine', 'tang blue', 'lime', 'antique white', 'tea green', 'chili red', 'coyote', 'parchment', 'turquoise',
    'verdigris', 'dark spring green', 'ruddy blue', 'carnation pink', 'tomato', 'ucla blue', 'hot pink', 'dark green',
    'tangelo', 'carolina blue', 'beige', 'bistre', 'pale azure', 'medium slate blue', 'tyrian purple', 'amaranth pink',
    'taupe gray', 'coffee', 'lavender blush', 'resolution blue', 'byzantine blue', 'magnolia', 'oxford blue',
    'cosmic latte', 'cream', 'lavender pink', 'air superiority blue', 'cal poly green', 'mint', 'cool gray', 'camel',
    'liver', 'maize', 'snow', 'rosy brown', 'lawn green', 'raspberry', 'folly', 'citron', 'orange peel',
    'anti-flash white', 'tawny', 'rust', 'china rose', 'aqua', 'giants orange', 'yinmn blue', 'blood red', 'khaki',
    'vista blue', 'golden gate bridge', 'rojo', 'arylide yellow', 'ebony', 'bittersweet', 'madder', 'cyan',
    'slate gray', 'lemon chiffon', 'shamrock green', 'ut orange', 'denim', 'gray', 'poppy', 'air force blue',
    'powder blue', 'uranian blue', 'french violet', 'salmon pink', 'turkey red', 'rose quartz', 'desert sand',
    'rebecca purple', 'international klein blue', 'aureolin', 'grape', 'midnight green', 'brandeis blue', 'wheat',
    'pearl', 'sage', 'safety orange', 'office green', 'rosewood', "hooker's green", 'light blue', 'scarlet', 'black',
    'light yellow', 'reseda green', 'magenta', 'tufts blue', 'smoky black', 'finn', 'green', 'hollywood cerise',
    "payne's gray", 'malachite', 'lapis lazuli', 'azul', 'tropical indigo', 'alice blue', 'eminence', 'tiffany blue',
    'zaffre', 'violet blue', "screamin'green", 'rose pink', 'off red', 'celestial blue', 'tan', 'pakistan green',
    'amaranth', 'dogwood rose', 'bittersweet shimmer', 'wenge', 'pomp and power', 'argentinian blue', 'flame',
    'maya blue', 'fire engine red', 'mikado yellow', 'quinacridone magenta', 'blue violet', 'mauve', 'telemagenta',
    'royal purple', 'rose pompadour', 'eggshell', 'aerospace orange', 'white smoke', 'fuchsia rose', 'vanilla', 'coral',
    'ivory', 'citrine', 'tangerine', 'eerie black', 'jet', 'navajo white', 'majorelle blue', 'british racing green',
    'azure', 'iris', 'navy blue', 'indigo dye', 'celtic blue', 'shocking pink', 'red', 'rose red', 'yellow green',
    'amethyst', 'pigment green', 'dark violet', 'gunmetal', 'puce', 'ecru', 'cornell red', 'russet',
    'school bus yellow', 'steel blue', 'white', 'silver', 'fulvous', 'mantis', 'chocolate cosmos', 'dark moss green',
    'buff', 'pink lavender', 'tea rose', 'champagne pink', 'sgbus green', 'rufous', 'black olive', 'heliotrope',
    'cherry blossom pink', 'claret', 'dark pastel green', 'candy apple red', 'brown', 'seashell', 'outer space',
    'electric indigo', 'harvest gold', 'burnt umber', 'federal blue', 'silver lake blue', 'canary', 'imperial red',
    'dark cyan', 'ghost white', 'teal', 'mimi pink', 'byzantium', 'ash gray', 'cinnabar', 'moss green', 'crimson',
    'non photo blue', 'hunter green', 'process cyan', 'mulberry', 'bleu de france', 'japanese violet', 'misty rose',
    'pink', 'dartmouth green', 'steel pink', 'purpureus', 'bice blue', 'dark slate gray', 'viridian', 'aquamarine',
    'rose bonbon', 'café noir', 'mahogany', 'celeste', 'rose ebony', 'neon green', 'tiger’s eye', 'columbia blue',
    'walnut brown', 'xanthous', 'charcoal', 'falu red', 'duke blue', 'penn blue', 'ultra violet', 'black bean', 'onyx',
    'persian red', 'beaver', 'lime green', 'jasmine', 'murrey', 'india green', 'nyanza', 'hunyadi yellow', 'apricot',
    'mint cream', 'polynesian blue', 'light red', 'french blue', 'jade', 'indian red', 'butterscotch', 'persian pink',
    'ice blue', 'dark purple', 'salmon', 'penn red', 'cornsilk', 'champagne', 'palatinate blue', 'fire brick',
    'sky blue', 'red violet', 'tekhelet', 'mauveine', 'taupe', 'zomp', 'pacific cyan', 'persimmon', 'electric violet',
    'spring green', 'lemon lime', 'persian orange', 'gold', 'bronze', 'olivine', 'skobeloff', 'jordy blue',
    'baby powder', 'seasalt', "davy's gray", 'fairy tale', 'keppel', 'bone', 'violet', 'isabelline', 'chocolate',
    'alabaster', 'cocoa brown', 'garnet', 'almond', 'russian violet', 'persian green', 'floral white', 'dark goldenrod',
    'cyclamen', 'satin sheen gold', 'plum', 'prussian blue', 'periwinkle', 'ou crimson', 'united nations blue',
    'saffron', 'selective yellow', 'bondi blue', 'coral pink', 'field drab', 'feldgrau', 'rich black', 'picton blue',
    'magenta dye', 'persian blue', 'glaucous', 'golden brown', 'savoy blue', 'van dyke', 'light coral', 'jungle green',
    'spring bud', 'pear', 'purple pizzazz', 'baby blue', 'lilac', 'alloy orange', 'baker-miller pink', 'indigo',
    'slate blue', 'sinopia', 'robin egg blue', 'deep pink', 'true blue', 'light cyan', 'midnight blue', 'sky magenta',
    'orchid pink', 'light sea green', 'orchid', 'bright green', 'light sky blue', 'carmine', 'pistachio', 'old gold',
    'orange', 'maroon', 'dark red', 'cordovan', 'french rose', 'spanish orange', 'fern green', 'old rose', 'yellow',
    'phlox', 'ochre', 'risd blue', 'english violet', 'fawn', 'fuchsia', 'brown sugar', 'royal blue', 'chefchaouen blue',
    'medium blue', 'tickle me pink', 'pale dogwood', 'redwood', 'capri', 'timberwolf', 'cerulean', 'honeydew', 'sienna',
    'licorice', 'mexican pink', 'cobalt blue', 'dodger blue', 'atomic tangerine', 'electric purple', 'light green',
    'dim gray', 'berkeley blue', 'purple', 'hot magenta', 'neon blue', 'chartreuse', 'chamoisee', 'earth yellow',
    'fandango', 'ultramarine', 'light orange', 'cinereous', 'sandy brown', 'burgundy', 'raisin black', 'avocado',
    'magenta haze', 'chinese violet', 'celadon', 'bright pink', 'mustard', 'blue green', 'cantaloupe melon', 'caramel',
    'green blue', 'electric blue', 'carrot orange', 'rusty red', 'drab dark brown', 'bole', 'african violet',
    'princeton orange', 'sunset', 'goldenrod', 'syracuse red orange', 'mindaro', 'marian blue', 'lavender', 'jonquil',
    'erin', 'peach yellow', 'harlequin', 'amber', 'jasper', 'rose', 'moonstone', 'vermilion', 'rose taupe',
    'apple green', 'cardinal', 'chestnut', 'pumpkin', 'flax', 'blue', 'razzle dazzle rose', 'veronica', 'wisteria',
    'sepia', 'asparagus', 'emerald', 'pale purple', 'kelly green', 'castleton green', 'burnt orange', 'dark magenta',
    'blue gray', 'papaya whip', 'aero', 'persian rose', 'delft blue', 'battleship gray', 'french mauve', 'green yellow',
    'thistle', 'copper', 'phthalo blue', 'palatinate', 'raspberry rose', 'forest green', 'mardi gras', 'gamboge',
    'french gray', 'raw umber', 'vivid sky blue', 'engineering orange', 'night', 'icterine', 'myrtle green', 'straw',
    'barn red', 'space cadet', 'cadet gray', 'sapphire', 'cambridge blue', 'dark orange', 'cornflower blue',
    'seal brown', 'thulian pink', 'old lace', 'kobicha', 'caribbean current', 'olive', 'sunglow', 'coquelicot',
    'amaranth purple', 'naples yellow', 'persian indigo', 'eggplant', 'mint green', 'burnt sienna', 'mountbatten pink',
    'chrysler blue', 'dutch white', 'sea green', 'dun', 'razzmatazz', 'linen', 'cerise', 'egyptian blue', 'yale blue',
    'blush', 'honolulu blue'
]

# IDBlindingTool: Global Variables
user_labels_list = []
ids_list = []
id_label_pairs = {}


# IDBlindingTool: Functions
def validate_inserted_ids():
    global ids_list
    i_ids = str(inputed_ids.get())
    if i_ids:
        duplicate_i_ids = len(list(i_ids.split(','))) != len(list(set(i_ids.split(','))))
        if duplicate_i_ids:
            messagebox.showerror('Error', 'Input has duplicates.\nPlease try again')
        else:
            while i_ids.endswith(','):
                i_ids = i_ids[: -1]
            while i_ids.startswith(','):
                i_ids = i_ids[1:]
            while ',,' in i_ids:
                i_ids = i_ids.replace(',,', ',')
            if ' ,' in i_ids:
                i_ids.replace(' ,', ',')
            if ', ' in i_ids:
                i_ids.replace(', ', ',')

            len_og_ids_list = len(ids_list)
            ids_list += list(set(i_ids.split(',')) - set(ids_list))  # Add only unique ids
            root.after(10, check_random_condition)
            ttkb.Label(mode1_frame, text=f'Currently, there are {len(ids_list)} IDs').grid(row=5, column=0)
            messagebox.showinfo('Success', f'{len(ids_list) - len_og_ids_list} IDs were added\n'
                                           f'In total, {len(ids_list)} IDs were successefuly inserted')


def validate_imported_ids():
    global ids_list
    ids_filename = filedialog.askopenfilename(
        initialdir=home_directory,
        title='Select File',
        filetypes=(('Supported Files: .txt;.csv;.xlsx;.xls', '*.txt;*.csv;*.xlsx;*.xls'), ('All files', '*.*')))
    if ids_filename:
        if ids_filename.endswith('.txt'):
            with open(ids_filename, 'r') as txt_file:
                ids_file_content = txt_file.read()
                if '\n' in ids_file_content:
                    ids_file_content = ids_file_content.replace('\n', ',')

        elif ids_filename.endswith('.csv'):
            with open(ids_filename, 'r') as csv_file:
                reader = csv.reader(csv_file)
                ids_file_content = ','.join([','.join(row) for row in reader])

        elif ids_filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(ids_filename)
            ws = wb[wb.sheetnames[0]]
            ids_file_content = ','.join([','.join(map(str, row)) for row in ws.iter_rows(values_only=True)])

        elif ids_filename.endswith('.xls'):
            wb = xlrd.open_workbook(ids_filename)
            sheet = wb.sheet_by_index(0)
            ids_file_content = ','.join([','.join(map(str, sheet.row_values(row))) for row in range(sheet.nrows)])

        else:
            ids_file_content = None
            messagebox.showerror('Error', 'Wrong Format.\nPlease try again')

        if ids_file_content is not None:
            while ids_file_content.endswith(','):
                ids_file_content = ids_file_content[: -1]
            while ids_file_content.startswith(','):
                ids_file_content = ids_file_content[1:]
            while ',,' in ids_file_content:
                ids_file_content = ids_file_content.replace(',,', ',')
            if ' ,' in ids_file_content:
                ids_file_content.replace(' ,', ',')
            if ', ' in ids_file_content:
                ids_file_content.replace(', ', ',')

            len_og_ids_list = len(ids_list)
            ids_list += list(set(ids_file_content.split(',')) - set(ids_list))
            root.after(100, check_random_condition)
            ttkb.Label(mode1_frame, text=f'Currently, there are {len(ids_list)} IDs').grid(row=5, column=0)
            messagebox.showinfo('Success', f'{len(ids_list) - len_og_ids_list} IDs were added\n'
                                           f'In total, {len(ids_list)} IDs were successefuly imported')


def validate_inserted_labels():
    global user_labels_list
    i_new_labels_list = str(inserted_labels.get())
    if i_new_labels_list:
        while i_new_labels_list.endswith(','):
            i_new_labels_list = i_new_labels_list[: -1]
        while i_new_labels_list.startswith(','):
            i_new_labels_list = i_new_labels_list[1:]
        while ',,' in i_new_labels_list:
            i_new_labels_list = i_new_labels_list.replace(',,', ',')
        if ' ,' in i_new_labels_list:
            i_new_labels_list.replace(' ,', ',')
        if ', ' in i_new_labels_list:
            i_new_labels_list.replace(', ', ',')

        len_og_user_labels_list = len(user_labels_list)
        user_labels_list += list(set(i_new_labels_list.split(',')) - set(user_labels_list))  # Add only unique labels
        root.after(100, check_random_condition)
        (ttkb.Label(mode1_frame, text=f'Currently, there are {len(user_labels_list)} labels')
         .grid(row=10, column=0, pady=(0, 10)))
        messagebox.showinfo('Success', f'{len(user_labels_list) - len_og_user_labels_list}labels were added\n'
                                       f'In total, {len(user_labels_list)} labels were successfuly imported')


def validate_imported_labels():
    global user_labels_list
    labels_filename = filedialog.askopenfilename(
        initialdir=home_directory,
        title='Select File',
        filetypes=(('Supported Files: .txt;.csv;.xlsx;.xls', '*.txt;*.csv;*.xlsx;*.xls'), ('All files', '*.*')))
    if labels_filename:
        if labels_filename.endswith('.txt'):
            with open(labels_filename, 'r') as txt_file:
                labels_file_content = txt_file.read()
                if '\n' in labels_file_content:
                    labels_file_content = labels_file_content.replace('\n', ',')

        elif labels_filename.endswith('.csv'):
            with open(labels_filename, 'r') as csv_file:
                reader = csv.reader(csv_file)
                labels_file_content = ','.join([','.join(row) for row in reader])

        elif labels_filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(labels_filename)
            ws = wb[wb.sheetnames[0]]
            labels_file_content = ','.join([','.join(map(str, row)) for row in ws.iter_rows(values_only=True)])

        elif labels_filename.endswith('.xls'):
            wb = xlrd.open_workbook(labels_filename)
            sheet = wb.sheet_by_index(0)
            labels_file_content = ','.join([','.join(map(str, sheet.row_values(row))) for row in range(sheet.nrows)])

        else:
            labels_file_content = None
            messagebox.showerror('Error', 'Wrong Format.\nPlease try again')

        if labels_file_content is not None:
            while labels_file_content.endswith(','):
                labels_file_content = labels_file_content[: -1]
            while labels_file_content.startswith(','):
                labels_file_content = labels_file_content[1:]
            while ',,' in labels_file_content:
                labels_file_content = labels_file_content.replace(',,', ',')
            if ' ,' in labels_file_content:
                labels_file_content.replace(' ,', ',')
            if ', ' in labels_file_content:
                labels_file_content.replace(', ', ',')

            len_og_user_labels_list = len(user_labels_list)
            user_labels_list += list(set(labels_file_content.split(',')) - set(user_labels_list))
            root.after(100, check_random_condition)
            (ttkb.Label(mode1_frame, text=f'Currently, there are {len(user_labels_list)} labels')
             .grid(row=10, column=0, pady=(0, 10)))
            messagebox.showinfo('Success', f'{len(user_labels_list) - len_og_user_labels_list}labels were added\n'
                                           f'In total, {len(user_labels_list)} labels were successefuly imported')


def check_random_condition():
    if use_default_labels.get() == 0:
        number_labels = len(user_labels_list)
    else:
        number_labels = len(og_labels_list)

    if len(ids_list) > 0:
        if len(ids_list) <= number_labels:
            randomize_button.config(state='normal')
        else:
            randomize_button.config(state='disabled')

    mode1_frame.update_idletasks()
    frame_width = mode1_frame.winfo_reqwidth()
    frame_height = mode1_frame.winfo_reqheight()
    root.geometry(f"{frame_width}x{frame_height}")


def randomize():
    global id_label_pairs
    if use_default_labels.get() == 0:
        labels_list = user_labels_list
    else:
        labels_list = og_labels_list
    random_label_selection = random.sample(labels_list, len(ids_list))  # Randomly pick labels in sufficient amount
    id_label_pairs = dict(zip(ids_list, random_label_selection))  # Assign each unique ID to a label at random

    if id_label_pairs is not None:
        messagebox.showinfo('Success', 'IDs have been randomly assigned to a label')
    else:
        messagebox.showerror('Error', 'There was a problem in randomly assigning the IDs to labels')

    output_filepath = filedialog.asksaveasfilename(
        initialdir=home_directory,
        title='Save As',
        filetypes=(('.xlsx', '*.xlsx'), ('All files', '*.*'))
    )
    if output_filepath:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        data = [['ID', 'Label']] + [[id_val, label_val] for id_val, label_val in id_label_pairs.items()]
        for row in data:
            sheet.append(row)

        workbook.save(output_filepath + '.xlsx')

        rename_button.config(state='normal')


def load_id_label_pairs():
    global id_label_pairs
    id_label_pairs = {}
    ids_label_pairs_filename = filedialog.askopenfilename(
        initialdir=home_directory,
        title='Select File',
        filetypes=(('Supported Files: .csv;.xlsx;.xls', '*.csv;*.xlsx;*.xls'), ('All files', '*.*')))
    if ids_label_pairs_filename:
        header1 = None
        header2 = None

        if ids_label_pairs_filename.endswith('.csv'):
            with open(ids_label_pairs_filename, 'r') as csv_file:
                reader = csv.reader(csv_file)
                for i, row in enumerate(reader):
                    if file_with_header.get() and i == 0:
                        header1, header2 = row[0], row[1]
                        continue
                    id_val, label_val = row[0], row[1]
                    id_label_pairs[id_val] = label_val

        elif ids_label_pairs_filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(ids_label_pairs_filename)
            ws = wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(min_row=1 if file_with_header.get() == 0 else 0, values_only=True)):
                if file_with_header.get() == 0 and i == 0:
                    header1, header2 = row[0], row[1]
                    continue
                id_val, label_val = row[0], row[1]
                id_label_pairs[id_val] = label_val

        elif ids_label_pairs_filename.endswith('.xls'):
            wb = xlrd.open_workbook(ids_label_pairs_filename)
            sheet = wb.sheet_by_index(0)
            for row_index in range(1 if file_with_header.get() == 0 else 0, sheet.nrows):
                if file_with_header.get() == 0 and row_index == 0:
                    header1 = sheet.cell_value(row_index, 0)
                    header2 = sheet.cell_value(row_index, 1)
                    continue
                id_val = sheet.cell_value(row_index, 0)
                label_val = sheet.cell_value(row_index, 1)
                id_label_pairs[id_val] = label_val

        else:
            id_label_pairs = None
            messagebox.showerror('Error', 'Wrong Format.\nPlease try again')

        non_empty_keys = [key for key in id_label_pairs.keys() if key.strip() != '']
        non_empty_values = [value for value in id_label_pairs.values() if value.strip() != '']
        if len(non_empty_keys) != len(id_label_pairs.keys()):
            id_label_pairs = None
            messagebox.showinfo('Error',
                                f'There seems to be empty IDs')
        if len(non_empty_values) != len(id_label_pairs.values()):
            id_label_pairs = None
            messagebox.showinfo('Error',
                                f'There seems to be empty Labels')
        if len(non_empty_keys) != len(non_empty_values):
            id_label_pairs = None
            messagebox.showinfo('Error',
                                f'There seems to be a different number of IDs and Labels')

        if id_label_pairs is not None:
            rename_databases_button.config(state='normal')
            rename_files_button.config(state='normal')

            for key in id_label_pairs.keys():
                if key == header1:
                    del id_label_pairs[header1]
                if key == header2:
                    del id_label_pairs[header2]

            messagebox.showinfo('Success', f'In total, {len(id_label_pairs)} IDs Label pairs were imported')


def rename_databases():
    database_filename = filedialog.askopenfilename(
        initialdir=home_directory,
        title='Select File',
        filetypes=(('Supported Files: .xlsx', '*.xlsx'), ('All files', '*.*')))
    if database_filename.endswith('.xlsx'):
        if unblind_mode.get() == 1:
            pairs = {value: key for key, value in id_label_pairs.items()}
        else:
            pairs = id_label_pairs

        if int(sheet_to_rename.get()) == 0:
            sheet_to_rename.set(value=1)
        if int(column_to_rename.get()) == 0:
            column_to_rename.set(value=1)

        database_name = database_filename.split("\\")[-1]
        wb = openpyxl.load_workbook(database_filename)
        sheet = wb.worksheets[sheet_to_rename.get() - 1]
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_index, cell_value in enumerate(row, start=1):
                cell_value_str = str(cell_value)
                if cell_value_str in pairs:
                    new_value = pairs[cell_value_str]
                    sheet.cell(row=row_index, column=int(column_to_rename.get()), value=new_value)
        wb.save(database_filename)
        messagebox.showinfo('Success', f'{database_name} files was successefuly renamed')
    else:
        messagebox.showerror('Error', 'Wrong Format.\nPlease try again')


def rename_files():
    folder_to_rename = filedialog.askdirectory(initialdir=home_directory, title="Select Folder")
    sucess = 0
    if unblind_mode.get() == 1:
        pairs = {value: key for key, value in id_label_pairs.items()}
    else:
        pairs = id_label_pairs
    for file_name in os.listdir(folder_to_rename):
        file_path = os.path.join(folder_to_rename, file_name)
        if os.path.isfile(file_path):
            file_name_without_extension, file_extension = os.path.splitext(file_name)
            file_id = pairs.get(file_name_without_extension)
            if file_id is not None:
                new_file_name = f"{file_id}{file_extension}"
                new_file_path = os.path.join(folder_to_rename, new_file_name)
                os.rename(file_path, new_file_path)
                sucess += 1
    if sucess == 0:
        messagebox.showerror('Error', f'{sucess} files were renamed')
    else:
        messagebox.showinfo('Success', f'{sucess} files were successefuly renamed')


# tkinter: Functions
def load_input_new_labels():
    if use_default_labels.get() == 0:
        inserted_labels_entry.grid(row=7, column=0, padx=(10, 103), pady=(10, 5), sticky='nsew')
        validate_labels_button.grid(row=7, column=0, padx=(103, 10), pady=(10, 5), sticky='e')
        imported_labels_button.grid(row=8, column=0, padx=10, sticky='ew')
        format_label.grid(row=9, column=0, pady=(0, 10), sticky='n')
    else:
        inserted_labels_entry.grid_remove()
        validate_labels_button.grid_remove()
        imported_labels_button.grid_remove()
        format_label.grid_remove()
    root.after(100, check_random_condition)


def load_initial_frame():
    initial_frame.tkraise()
    initial_frame.grid_propagate(False)

    initial_frame.update_idletasks()
    frame_width = initial_frame.winfo_reqwidth()
    frame_height = initial_frame.winfo_reqheight()
    root.geometry(f"{frame_width}x{frame_height}")

    # Widget: Logo
    logo_img = ImageTk.PhotoImage(file=logo_img_path)
    logo_widget = ttkb.Label(initial_frame, image=logo_img)
    logo_widget.image = logo_img
    logo_widget.grid(row=0, column=0)

    ttkb.Label(initial_frame, text='Blinding Tool for Research', font='Aptos').grid(row=1, column=0)

    ttkb.Button(initial_frame, text='Randomly assign IDs to Labels',
                cursor='hand2', command=lambda: load_mode1()).grid(row=2, column=0, padx=10, pady=10, sticky='nsew')
    ttkb.Button(initial_frame, text='Rename Files using a Key',
                cursor='hand2', command=lambda: load_mode2()).grid(row=3, column=0, padx=10, pady=0, sticky='nsew')

    (ttkb.Button(initial_frame, text='GitHub Repository', style='dark', cursor='hand2', command=lambda: open_github())
     .grid(row=4, column=0, padx=10, pady=(100, 0), sticky='sw'))
    ttkb.Label(initial_frame, text='v.1.0.1', foreground='#868683').grid(row=4, column=0, padx=10, sticky='se')

    if use_default_labels.get() == 0:
        inserted_labels_entry.grid_remove()
        validate_labels_button.grid_remove()
        imported_labels_button.grid_remove()
        format_label.grid_remove()
        use_default_labels.set(1)


def load_info1():
    info1_window = tk.Toplevel(root)
    info1_window.iconbitmap(logo_icon_path)
    info1_window.title("Help - Assign IDs to Labels")
    info1_window.resizable(False, False)
    info1_frame = Frame(info1_window)
    info1_frame.grid(row=0, column=0, sticky='nsew')

    info1_canvas = Canvas(info1_frame, width=510, height=500)
    info1_canvas.grid(row=0, column=0, sticky='nsew')

    info1_scrollbar = tk.Scrollbar(info1_frame, orient=VERTICAL, command=info1_canvas.yview)
    info1_scrollbar.grid(row=0, column=1, rowspan=17, sticky='ns')

    info1_canvas.configure(yscrollcommand=info1_scrollbar.set)
    info1_canvas.bind('<Configure>', lambda e: info1_canvas.configure(scrollregion=info1_canvas.bbox('all')))

    info1_frame2 = Frame(info1_canvas, width=200, height=200)

    info_default_label = ttkb.Label(info1_frame2, text='Default Labels')
    info_default_separator = ttkb.Separator(info1_frame2, orient='horizontal')
    info_default_label1 = ttkb.Label(info1_frame2, text='The default labels are 509 color names')

    info_default_label.grid(row=1, column=0, columnspan=2, padx=10, pady=(10, 5), sticky='w')
    info_default_separator.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_default_label1.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky='w')

    info_insert_label = ttkb.Label(info1_frame2, text='Insert IDs and Labels')
    info_insert_separator = ttkb.Separator(info1_frame2, orient='horizontal')
    info_insert_delimiter = ttkb.Label(info1_frame2, text='Commas (,) as delimiters! Do not use in IDs or Labels.')
    info_insert_format1_label = ttkb.Label(info1_frame2, text='Case and Space Sensitive')
    info_insert_format2_label = ttkb.Label(info1_frame2, text='No leading or trailing spaces')
    info_insert_format3_label = ttkb.Label(info1_frame2, text='Requeried Format:')
    info_insert_format_one_value = ttkb.StringVar(value='text')
    info_insert_format_one_entry = tk.Entry(info1_frame2, textvariable=info_insert_format_one_value, state='readonly')
    info_insert_format_one_label = ttkb.Label(info1_frame2, text='Insert one by one')
    info_insert_format_multiple_value = ttkb.StringVar(value='text1,text2,[...],textn')
    info_insert_format_multiple_entry = tk.Entry(info1_frame2,
                                                 textvariable=info_insert_format_multiple_value, state='readonly')
    info_insert_format_multiple_label = ttkb.Label(info1_frame2, text='Insert a list')

    info_insert_label.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_insert_separator.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_insert_delimiter.grid(row=4, column=0, columnspan=2, padx=10, pady=(5, 0), sticky='w')
    info_insert_format1_label.grid(row=5, column=0, columnspan=2, padx=10, pady=0, sticky='w')
    info_insert_format2_label.grid(row=6, column=0, columnspan=2, padx=10, pady=0, sticky='w')
    info_insert_format3_label.grid(row=7, column=0, columnspan=2, padx=10, pady=(0, 5), sticky='w')
    info_insert_format_one_entry.grid(row=8, column=0, padx=10, pady=5, sticky='nsew')
    info_insert_format_one_label.grid(row=8, column=1, pady=5, sticky='w')
    info_insert_format_multiple_entry.grid(row=9, column=0, padx=10, pady=5, sticky='nsew')
    info_insert_format_multiple_label.grid(row=9, column=1, pady=5, sticky='w')

    info_import_label = ttkb.Label(info1_frame2, text='Import IDs and Labels')
    info_import_separator = ttkb.Separator(info1_frame2, orient='horizontal')
    info_import_delimiter = ttkb.Label(info1_frame2, text='Commas (,) as delimiters! Do not use in IDs or Labels.')
    info_import_label1 = ttkb.Label(info1_frame2, text='IDs can be imported from .xlsx, .xls, .csv, and .txt files')
    info_import_label2 = ttkb.Label(info1_frame2, text='Here is their required format:')
    import_formats_img = ImageTk.PhotoImage(file=import_formats_img_path)
    import_formats_widget = ttkb.Label(info1_frame2, image=import_formats_img)
    import_formats_widget.image = import_formats_img
    import_label = ttkb.Label(info1_frame2, text='Separated by rows or columns\n'
                                                 'Only text to import - no header\n'
                                                 'First sheet\n')

    info_import_label.grid(row=10, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_import_separator.grid(row=10, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_import_delimiter.grid(row=11, column=0, columnspan=2, padx=10, pady=(5, 0), sticky='w')
    info_import_label1.grid(row=12, column=0, columnspan=2, padx=10, pady=0, sticky='w')
    info_import_label2.grid(row=13, column=0, columnspan=2, padx=10, pady=(0, 5), sticky='w')
    import_formats_widget.grid(row=14, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    import_label.grid(row=14, column=1, padx=(190, 10), pady=5, sticky='w')

    info_key_label = ttkb.Label(info1_frame2, text='Create Key')
    info_key_separator = ttkb.Separator(info1_frame2, orient='horizontal')
    info_randomize_button = ttkb.Button(info1_frame2, text='Create Key', style='success')
    info_randomize_button_label1 = ttkb.Label(info1_frame2,
                                              text='Only active once there are enough labels to rename the loaded IDs')
    info_randomize_button_label2 = ttkb.Label(info1_frame2,
                                              text='Randomly assign unique IDs to distinct Labels and save the Key')

    info_key_label.grid(row=15, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_key_separator.grid(row=15, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_randomize_button_label1.grid(row=16, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_randomize_button.grid(row=17, column=0, padx=10, pady=5, sticky='nsew')
    info_randomize_button_label2.grid(row=17, column=1, padx=(0, 10), pady=5, sticky='w')

    info_rename_key_label = ttkb.Label(info1_frame2, text='Rename Files')
    info_rename_separator = ttkb.Separator(info1_frame2, orient='horizontal')
    info_rename_button_label1 = ttkb.Label(info1_frame2,
                                           text='Only active once there are enough labels to rename the loaded IDs')
    info_rename_button = ttkb.Button(info1_frame2, text='Rename', style='success')
    info_rename_button_label = ttkb.Label(info1_frame2,
                                          text='Load the Rename page. The loaded IDs and Labels are not lost')

    info_rename_key_label.grid(row=18, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_rename_separator.grid(row=18, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_rename_button_label1.grid(row=19, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_rename_button.grid(row=20, column=0, padx=10, pady=5, sticky='nsew')
    info_rename_button_label.grid(row=20, column=1, pady=5, sticky='w')

    info_home_label1 = ttkb.Label(info1_frame2, text='Return Home')
    info_home_separator1 = ttkb.Separator(info1_frame2, orient='horizontal')
    info_return_button1 = ttkb.Button(info1_frame2, text='HOME', style='dark')
    info_return_button_label1 = ttkb.Label(info1_frame2,
                                           text='Return to the initial page. The loaded IDs and Labels are not lost')

    info_home_label1.grid(row=21, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    info_home_separator1.grid(row=21, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_return_button1.grid(row=22, column=0, padx=10, pady=(5, 10), sticky='nsew')
    info_return_button_label1.grid(row=22, column=1, pady=(5, 10), sticky='w')

    info1_canvas.create_window((0, 0), window=info1_frame2, anchor='nw')


def load_mode1():
    mode1_frame.tkraise()

    # Widget: Logo
    logo_img = ImageTk.PhotoImage(file=logo_img_path)
    logo_widget = ttkb.Label(mode1_frame, image=logo_img)
    logo_widget.image = logo_img
    logo_widget.grid(row=0, column=0)

    ttkb.Label(mode1_frame, text='Randomly assign IDs to Labels',
               font='Aptos').grid(row=1, column=0, padx=10, sticky='n')

    # IDs Section
    inputed_ids_entry = (ttkb.Entry(mode1_frame, textvariable=inputed_ids)
                         .grid(row=2, column=0, padx=(10, 103), pady=(10, 5), sticky='nsew'))
    validate_ids_button = (ttkb.Button(mode1_frame, text='Insert IDs', cursor='hand2',
                                       command=lambda: validate_inserted_ids())
                           .grid(row=2, column=0, padx=(200, 10), pady=(10, 5), sticky='ew'))
    import_ids_button = (ttkb.Button(mode1_frame, text='Import IDs', cursor='hand2',
                                     command=lambda: validate_imported_ids())
                         .grid(row=3, column=0, padx=10, sticky='nsew'))
    ttkb.Label(mode1_frame, text='.txt, .csv, .xlsx, .xls').grid(row=4, column=0)

    # Label Section
    ttkb.Label(mode1_frame, text='Use Default Labels').grid(row=6, column=0, padx=(0, 30))

    use_default_labels_check = (ttkb.Checkbutton(mode1_frame, style='round-toggle',
                                                 variable=use_default_labels, onvalue=1, offvalue=0,
                                                 command=lambda: load_input_new_labels())
                                .grid(row=6, column=0, padx=(110, 0), pady=10))

    global inserted_labels_entry, validate_labels_button, imported_labels_button, format_label
    inserted_labels_entry = ttkb.Entry(mode1_frame, textvariable=inserted_labels)
    validate_labels_button = ttkb.Button(mode1_frame, text='Insert Labels', cursor='hand2',
                                         command=lambda: validate_inserted_labels())
    imported_labels_button = ttkb.Button(mode1_frame, text='Import Labels', cursor='hand2',
                                         command=lambda: validate_imported_labels())
    format_label = ttkb.Label(mode1_frame, text='.txt, .csv, .xlsx, .xls')

    # Separator
    ttkb.Separator(mode1_frame, orient='horizontal').grid(row=11, column=0, padx=10, sticky='ew')

    # Randomize
    global randomize_button
    randomize_button = ttkb.Button(mode1_frame, text='Create Key', style='success', state='disabled', cursor='hand2',
                                   command=lambda: randomize())
    randomize_button.grid(row=12, column=0, padx=10, pady=10, sticky='ew')

    # Rename Section
    global rename_button
    rename_button = ttkb.Button(mode1_frame, text='Rename', style='success', state='disabled', cursor='hand2',
                                command=lambda: load_mode2())
    rename_button.grid(row=13, column=0, padx=10, sticky='ew')

    # Home
    return_button = ttkb.Button(mode1_frame, text='HOME', style='dark', cursor='hand2',
                                command=lambda: load_initial_frame())
    return_button.grid(row=14, column=0, padx=(10, 152), pady=10, sticky='ew')

    # Info
    info_1_button = ttkb.Button(mode1_frame, text='HELP', style='dark', cursor='hand2', command=lambda: load_info1())
    info_1_button.grid(row=14, column=0, padx=(152, 10), pady=10, sticky='ew')

    mode1_frame.update_idletasks()
    frame_width = mode1_frame.winfo_reqwidth()
    frame_height = mode1_frame.winfo_reqheight()
    root.geometry(f"{frame_width}x{frame_height}")


def load_info2():
    info2_window = tk.Toplevel(root)
    info2_window.iconbitmap(logo_icon_path)
    info2_window.title("Help - Rename Files")
    info2_window.resizable(False, False)
    info2_frame = Frame(info2_window)
    info2_frame.grid(row=0, column=0, sticky='nsew')

    info2_canvas = Canvas(info2_frame, width=550, height=500)
    info2_canvas.grid(row=0, column=0, sticky='nsew')

    info2_scrollbar = tk.Scrollbar(info2_frame, orient=VERTICAL, command=info2_canvas.yview)
    info2_scrollbar.grid(row=0, column=1, sticky='ns')

    info2_canvas.configure(yscrollcommand=info2_scrollbar.set)
    info2_canvas.bind('<Configure>', lambda e: info2_canvas.configure(scrollregion=info2_canvas.bbox('all')))

    info2_frame2 = Frame(info2_canvas, width=200, height=200)

    info_load_key_label = ttkb.Label(info2_frame2, text='Load Key')
    info_load_key_separator = ttkb.Separator(info2_frame2, orient='horizontal')
    info_validate_ids_button = ttkb.Button(info2_frame2, text='Load Key')
    info_validate_ids_button_label = ttkb.Label(info2_frame2, text='Insert a key in .xlsx format')
    info_input_format_label1 = ttkb.Label(info2_frame2,
                                          text='IDs must be in the first column\n'
                                               'Labels must be in the second column\n'
                                               'The file may have an header,\n'
                                               'the slider should relfect this property of the file')
    info_input_format_label2 = ttkb.Label(info2_frame2, text='This should be the format:')
    key_img = ImageTk.PhotoImage(file=key_img_path)
    key_widget = ttkb.Label(info2_frame2, image=key_img)
    key_widget.image = key_img
    info_rename_label = ttkb.Label(info2_frame2,
                                   text='Once a Key was loaded, the Rename Database and Rename Files become active')

    info_load_key_label.grid(row=1, column=0, columnspan=2, padx=10, pady=(10, 5), sticky='w')
    info_load_key_separator.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_validate_ids_button.grid(row=2, column=0, padx=10, pady=5, sticky='nsew')
    info_validate_ids_button_label.grid(row=2, column=1, padx=(0, 10), pady=5, sticky='w')
    info_input_format_label1.grid(row=4, column=1, padx=(100, 10), pady=(5, 0), sticky='w')
    info_input_format_label2.grid(row=3, column=0, columnspan=2, padx=10, pady=(0, 5), sticky='w')
    key_widget.grid(row=4, column=0, columnspan=2, sticky='w')

    info_database_label = ttkb.Label(info2_frame2, text='Rename Databases')
    info_database_separator = ttkb.Separator(info2_frame2, orient='horizontal')
    info_rename_databases_label = ttkb.Label(info2_frame2,
                                             text='Only a column in a sheet in one file can be renamed at a time')
    info_sheet_label = ttkb.Label(info2_frame2, text='Sheet: Specify which sheet of the file to change\n'
                                                     'Column: Specify which column in the sheet to change')
    info_rename_databases_button = ttkb.Button(info2_frame2, text='Rename Database', style='success')
    info_rename_databases_button_label = ttkb.Label(info2_frame2,
                                                    text='Choose an .xlsx file to modify cells '
                                                         'within the designated column and sheet')
    db_img = ImageTk.PhotoImage(file=db_img_path)
    db_widget = ttkb.Label(info2_frame2, image=db_img)
    db_widget.image = db_img

    info_database_label.grid(row=7, column=0, columnspan=2, padx=10, pady=(10, 5), sticky='w')
    info_database_separator.grid(row=7, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_rename_databases_label.grid(row=9, column=0, columnspan=2, padx=10, pady=(5, 0), sticky='w')
    info_sheet_label.grid(row=10, column=1, padx=(100, 10), sticky='w')
    info_rename_databases_button.grid(row=8, column=0, padx=10, pady=5, sticky='nsew')
    info_rename_databases_button_label.grid(row=8, column=1, padx=(0, 10), pady=5, sticky='w')
    db_widget.grid(row=10, column=0, columnspan=2, sticky='w')

    info_files_label = ttkb.Label(info2_frame2, text='Rename Files')
    info_files_separator = ttkb.Separator(info2_frame2, orient='horizontal')
    info_rename_files_button = ttkb.Button(info2_frame2, text='Rename Files', style='success')
    info_rename_files_button_label = ttkb.Label(info2_frame2,
                                                text='Select a folder to rename the files based on the loaded key')
    info_rename_files_label = ttkb.Label(info2_frame2, text='Filenames need to exactly match the key')
    files_img = ImageTk.PhotoImage(file=files_img_path)
    files_widget = ttkb.Label(info2_frame2, image=files_img)
    files_widget.image = files_img

    info_files_label.grid(row=13, column=0, columnspan=2, padx=10, pady=(10, 5), sticky='w')
    info_files_separator.grid(row=13, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_rename_files_button.grid(row=14, column=0, padx=10, pady=5, sticky='nsew')
    info_rename_files_button_label.grid(row=14, column=1, padx=(0, 10), pady=5, sticky='w')
    files_widget.grid(row=15, column=0, columnspan=2, sticky='w')
    info_rename_files_label.grid(row=15, column=1, padx=(160, 10), sticky='w')

    info_home_label = ttkb.Label(info2_frame2, text='Return Home')
    info_home_separator = ttkb.Separator(info2_frame2, orient='horizontal')
    info_return_button = ttkb.Button(info2_frame2, text='HOME', style='dark')
    info_return_button_label = ttkb.Label(info2_frame2, text='Return to the initial page. The Key stays loaded')

    info_home_label.grid(row=16, column=0, columnspan=2, padx=10, pady=(10, 5), sticky='w')
    info_home_separator.grid(row=16, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_return_button.grid(row=17, column=0, padx=10, pady=5, sticky='nsew')
    info_return_button_label.grid(row=17, column=1, padx=(0, 10), pady=5, sticky='w')

    ideal_db_label = ttkb.Label(info2_frame2, text='How to Completely Rename a Database')
    info_ideal_db_separator = ttkb.Separator(info2_frame2, orient='horizontal')
    info_ideal_db_label = ttkb.Label(info2_frame2, text='To Rename IDs and Categorical Variables:\n'
                                                        '1. Create a Key for the IDs categorical variables\n'
                                                        '2. Use these Keys to rename the specif columns\n'
                                                        'Example:')
    ideal_db_img = ImageTk.PhotoImage(file=ideal_db_img_path)
    ideal_db_widget = ttkb.Label(info2_frame2, image=ideal_db_img)
    ideal_db_widget.image = ideal_db_img

    ideal_db_label.grid(row=18, column=0, columnspan=2, padx=10, pady=(10, 5), sticky='w')
    info_ideal_db_separator.grid(row=18, column=0, columnspan=2, padx=10, pady=5, sticky='ew' + 's')
    info_ideal_db_label.grid(row=19, column=0, columnspan=2, padx=10, pady=5, sticky='w')
    ideal_db_widget.grid(row=20, column=0, columnspan=2, padx=10, pady=(5, 10), sticky='w')

    info2_canvas.create_window((0, 0), window=info2_frame2, anchor='nw')


def load_mode2():
    mode2_frame.tkraise()

    # Widget Logo
    logo_img = ImageTk.PhotoImage(file=logo_img_path)
    logo_widget = ttkb.Label(mode2_frame, image=logo_img)
    logo_widget.image = logo_img
    logo_widget.grid(row=0, column=0)

    ttkb.Label(mode2_frame, text='Rename Files using a Key',
               font='Aptos').grid(row=1, column=0, padx=10, sticky='n')

    # Unblind
    ttkb.Label(mode2_frame, text='Blind').grid(row=2, column=0, padx=(0, 70))
    ttkb.Label(mode2_frame, text='Unblind').grid(row=2, column=0, padx=(70, 0))
    unblind_mode_check = (ttkb.Checkbutton(mode2_frame, style='primary, round-toggle',
                                           variable=unblind_mode, onvalue=1, offvalue=0)
                          .grid(row=2, column=0, padx=(0, 0), pady=10))

    # Load ID-Label pairs
    validate_ids_button = (ttkb.Button(mode2_frame, text='Load Key', cursor='hand2',
                                       command=lambda: load_id_label_pairs())
                           .grid(row=3, column=0, padx=10, pady=10, sticky='nsew'))

    # Format Input:
    ttkb.Label(mode2_frame, text='First Row is an Header').grid(row=4, column=0, padx=(0, 55), pady=(0, 2))
    file_with_header_check = (ttkb.Checkbutton(mode2_frame, style='primary, round-toggle',
                                               variable=file_with_header, onvalue=0, offvalue=1)
                              .grid(row=4, column=0, padx=(100, 0)))

    # Separator
    ttkb.Separator(mode2_frame, orient='horizontal').grid(row=5, column=0, pady=(10, 0), sticky='ew')

    # Rename Databases
    global rename_databases_button
    rename_databases_button = ttkb.Button(mode2_frame, text='Rename Database', style='success',
                                          state='disabled', cursor='hand2', command=lambda: rename_databases())
    rename_databases_button.grid(row=6, column=0, rowspan=2, padx=(120, 10), pady=(10, 0), sticky='nsew')
    ttkb.Label(mode2_frame, text='Sheet:').grid(row=6, column=0, padx=(10, 0), pady=(10, 0), sticky='w')
    sheet_entry = (ttkb.Entry(mode2_frame, textvariable=sheet_to_rename, width=5)
                   .grid(row=6, column=0, padx=(65, 0), pady=(10, 0), sticky='w'))
    ttkb.Label(mode2_frame, text='Column:').grid(row=7, column=0, padx=(10, 0), pady=(10, 0), sticky='w')
    sheet_entry = (ttkb.Entry(mode2_frame, textvariable=column_to_rename, width=5)
                   .grid(row=7, column=0, padx=(65, 0), pady=(10, 0), sticky='w'))

    # Rename Files
    global rename_files_button
    rename_files_button = ttkb.Button(mode2_frame, text='Rename Files', style='success',
                                      state='disabled', cursor='hand2', command=lambda: rename_files())
    rename_files_button.grid(row=12, column=0, padx=10, pady=(10, 0), sticky='ew')

    # HOME
    return_button = ttkb.Button(mode2_frame, text='HOME', style='dark', cursor='hand2',
                                command=lambda: load_initial_frame())
    return_button.grid(row=13, column=0, padx=(10, 152), pady=10, sticky='ew')

    # Info
    info_2_button = ttkb.Button(mode2_frame, text='HELP', style='dark', cursor='hand2', command=lambda: load_info2())
    info_2_button.grid(row=13, column=0, padx=(152, 10), pady=10, sticky='ew')

    mode2_frame.update_idletasks()
    frame_width = mode2_frame.winfo_reqwidth()
    frame_height = mode2_frame.winfo_reqheight()
    root.geometry(f"{frame_width}x{frame_height}")


# tkinter Setup
root = Tk()
style = ttkb.Style(theme='lumen')
root.iconbitmap(logo_icon_path)
root.title('ID Blinding Tool')
root.resizable(False, False)

# tkinter Frames
initial_frame = ttkb.Frame(root, width=300, height=400)
mode1_frame = ttkb.Frame(root)
mode2_frame = ttkb.Frame(root)

# Tkinter: Global Variables
inputed_ids = StringVar()
imported_ids = StringVar()
use_default_labels = IntVar(value=1)
inserted_labels_entry = ttkb.Entry()
inserted_labels = StringVar()
validate_labels_button = ttkb.Button()
imported_labels_button = ttkb.Button()
format_label = ttkb.Label()
randomize_button = ttkb.Button()
file_with_header = IntVar(value=0)
unblind_mode = IntVar(value=0)
rename_button = ttkb.Button()
rename_databases_button = ttkb.Button()
sheet_to_rename = IntVar(value=1)
column_to_rename = IntVar(value=1)
rename_files_button = ttkb.Button()

for frame in (initial_frame, mode1_frame, mode2_frame):
    frame.grid(row=0, column=0, sticky='nsew')

# tkinter Load Initial Page
load_initial_frame()

# tkinter Run GUI
root.mainloop()
